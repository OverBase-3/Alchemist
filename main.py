import argparse
import logging
import sys
import time
import math
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import joblib
from openpyxl.styles import Font, PatternFill, Alignment

from bio_groups      import normalize_cell_type
from data_loader     import DataLoader
from feature_engineering import FeatureEngineer
from model_training  import ModelTrainer
from group_model     import GroupModelSystem
from media_optimizer import MediaOptimizer, MediaFormulation
from bio_constraints import (
    get_config, calculate_osmolarity, cdm_ceiling
)

logger = logging.getLogger(__name__)

TARGET_GROUP   = "E.coli"
TARGETS        = ["growth_rate", "max_cell_density"]
MODEL_PATH     = "results/alchemist_model.pkl"
RESULTS_EXCEL  = "results/alchemist_results.xlsx"

def _fmt_time(s: float) -> str:
    return f"{s:.1f}s" if s < 60 else f"{s/60:.1f}min"


def _setup_logging(log_level: str):
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    logging.basicConfig(
        level   = getattr(logging, log_level.upper()),
        format  = fmt,
        datefmt = "%H:%M:%S",
        handlers=[
            logging.FileHandler("alchemist.log", encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def _banner(title: str):
    w = 62
    print(f"\n{'═'*w}")
    print(f"  {title}")
    print(f"{'═'*w}")

def train_model(data_path: str):

    _banner("PASO 1 — Carga de datos")
    loader = DataLoader(data_path)
    exp, gc, cc = loader.load()
    logger.info(f"  {loader.get_summary()}")

    exp_ecoli = exp[exp["cell_type"].apply(normalize_cell_type) == TARGET_GROUP].copy()
    cc_ecoli  = cc[cc["medium_id"].isin(exp_ecoli["medium_id"].unique())].copy()

    n_medios = exp_ecoli["medium_id"].nunique()
    n_comps  = exp_ecoli["component"].nunique()
    logger.info(f"  {TARGET_GROUP}: {n_medios} medios | {n_comps} componentes únicos")

    if n_medios < 3:
        raise ValueError(f"Solo {n_medios} medios de {TARGET_GROUP}. Mínimo: 3.")

    _banner("PASO 2 — Feature engineering")
    fe = FeatureEngineer(targets=TARGETS, freq_threshold=0.20)
    X, y = fe.fit_transform(exp_ecoli, cc_ecoli)
    logger.info(f"  Matriz X: {X.shape[0]} muestras × {X.shape[1]} features")

    _banner("PASO 3 — Entrenamiento del modelo")
    trainer = ModelTrainer(targets=TARGETS)
    trainer.fit(X, y)
    best = trainer.best_model_name
    logger.info(f"  Mejor modelo: {best}")

    _banner("PASO 3b — Rendimiento within-group (E.coli)")
    try:
        results_df = trainer.get_results_table()
        logger.info(f"\n{results_df.to_string(index=False)}")
    except Exception:
        pass

    # Guardar modelo
    Path("results").mkdir(exist_ok=True)
    joblib.dump({
        "trainers":          {TARGET_GROUP: trainer},
        "feature_engineers": {TARGET_GROUP: fe},
        "exp_data_by_group": {TARGET_GROUP: exp_ecoli},
    }, MODEL_PATH)
    logger.info(f"  Modelo guardado: {MODEL_PATH}")

    return trainer, fe, exp_ecoli


def run_optimize(data_path: str, n_trials: int, n_top: int):
    t0 = time.time()

    trainer, fe, exp_ecoli = train_model(data_path)

    _banner("PASO 4 — Optimizacion de medios (E.coli)")
    cfg_env = get_config(TARGET_GROUP)["condiciones_entorno"]
    logger.info(f"  Hard bounds: pH [{cfg_env['pH']['min']}–{cfg_env['pH']['max']}]  "
                f"T [{cfg_env['temperature_C']['min']}–{cfg_env['temperature_C']['max']}°C]  "
                f"rpm [{cfg_env['agitation_rpm']['min']}–{cfg_env['agitation_rpm']['max']}]")

    optimizer = MediaOptimizer(trainer, fe, exp_ecoli, n_top=n_top)
    formulaciones = optimizer.optimize(n_trials=n_trials)
    print()
    for f in formulaciones[:5]:
        carriers_str = "yes" if float(f.conditions.get("carriers", 0)) >= 0.5 else "no"
        logger.info(
            f"  #{f.rank:2d} | CDM={f.predicted_density:.3f} g/L | "
            f"µ={f.predicted_growth:.4f} h⁻¹ | td={f.predicted_doubling_time:.2f}h | "
            f"Score={f.optimization_score:.4f} | bio_pen={f.bio_penalty:.3f} | "
            f"pH={f.conditions.get('pH',0):.2f} "
            f"T={f.conditions.get('temperature_C',0):.0f}°C "
            f"rpm={f.conditions.get('agitation_rpm',0):.0f} "
            f"carriers={carriers_str}"
        )
        osm, _ = calculate_osmolarity(f.composition)
        zona   = f.osm_detail.get("zona", "?") if f.osm_detail else "?"
        logger.info(f"       Osm: {osm:.0f} mOsm/kg  zona={zona}")
        for adv in f.advertencias:
            logger.info(f"       ⚠  [{adv.get('capa','')}] {adv.get('mensaje','')[:80]}")

    _banner("PASO 5 — Exportacion de resultados")
    filter_report = optimizer.get_filter_report(formulaciones)
    export_results_excel(formulaciones, filter_report, RESULTS_EXCEL)
    logger.info(f"  Resultados: {RESULTS_EXCEL}")
    logger.info(f"  Completado en {_fmt_time(time.time()-t0)}")


def export_results_excel(formulaciones: list, filter_df: pd.DataFrame,
                          output_path: str):
    """Genera el Excel de resultados con 5 hojas."""

    rows_media = []
    for f in formulaciones:
        conds = f.conditions
        osm, _ = calculate_osmolarity(f.composition)
        for comp, conc in f.composition.items():
            rows_media.append({
                "rank":               f.rank,
                "medium_id":          f"OPT_{f.rank:03d}",
                "cell_type":          f.cell_type,
                "component":          comp,
                "concentration":      conc,
                "unit":               "g/L",
                "growth_rate":        f.predicted_growth,
                "max_cell_density":   f.predicted_density,
                "doubling_time":      f.predicted_doubling_time,
                "temperature_C":      conds.get("temperature_C"),
                "pH":                 conds.get("pH"),
                "agitation_rpm":      conds.get("agitation_rpm"),
                "carriers":           "yes" if conds.get("carriers",0) >= 0.5 else "no",
                "optimization_score": f.optimization_score,
                "bio_penalty":        f.bio_penalty,
                "osmolarity_mOsm":    round(osm, 1),
                "pass":               f.pass_number,
                "method":             f.method,
                "n_advertencias":     len(f.advertencias),
            })

    df_media = pd.DataFrame(rows_media)
    rows_adv = []
    for f in formulaciones:
        for adv in f.advertencias:
            rows_adv.append({
                "rank":      f.rank,
                "medium_id": f"OPT_{f.rank:03d}",
                "capa":      adv.get("capa",""),
                "variable":  adv.get("variable", adv.get("componente","")),
                "mensaje":   adv.get("mensaje",""),
            })
    df_adv = pd.DataFrame(rows_adv) if rows_adv else pd.DataFrame(
        columns=["rank","medium_id","capa","variable","mensaje"])

    resumen_rows = []
    for f in formulaciones:
        osm, _ = calculate_osmolarity(f.composition)
        resumen_rows.append({
            "rank":             f.rank,
            "medium_id":        f"OPT_{f.rank:03d}",
            "µ_h":              f.predicted_growth,
            "td_h":             f.predicted_doubling_time,
            "CDM_g_L":          f.predicted_density,
            "score":            f.optimization_score,
            "bio_penalty":      f.bio_penalty,
            "osm_mOsm_kg":      round(osm, 1),
            "n_componentes":    f.n_components,
            "n_advertencias":   len(f.advertencias),
            "pass":             f.pass_number,
        })
    df_resumen = pd.DataFrame(resumen_rows)

    cfg = get_config(TARGET_GROUP)
    bio_rows = []
    env = cfg["condiciones_entorno"]
    bio_rows.append({"parametro": "pH minimo",          "valor": env["pH"]["min"],          "unidad": "", "fuente": "JSON bio_config"})
    bio_rows.append({"parametro": "pH maximo",          "valor": env["pH"]["max"],          "unidad": "", "fuente": "JSON bio_config"})
    bio_rows.append({"parametro": "pH optimo",          "valor": env["pH"]["optimo"],       "unidad": "", "fuente": "JSON bio_config"})
    bio_rows.append({"parametro": "T minima",           "valor": env["temperature_C"]["min"], "unidad": "°C", "fuente": "JSON bio_config"})
    bio_rows.append({"parametro": "T maxima",           "valor": env["temperature_C"]["max"], "unidad": "°C", "fuente": "JSON bio_config"})
    bio_rows.append({"parametro": "T optima",           "valor": env["temperature_C"]["optimo"], "unidad": "°C", "fuente": "JSON bio_config"})
    bio_rows.append({"parametro": "rpm minimo",         "valor": env["agitation_rpm"]["min"], "unidad": "rpm", "fuente": "JSON bio_config"})
    bio_rows.append({"parametro": "rpm maximo",         "valor": env["agitation_rpm"]["max"], "unidad": "rpm", "fuente": "JSON bio_config"})
    bio_rows.append({"parametro": "N fraccion biomasa (Pirt)", "valor": cfg["estequiometria_biomasa"]["N_fraction"], "unidad": "g N/g CDM", "fuente": "Pirt 1965 / Roels 1980"})
    bio_rows.append({"parametro": "Yx/s aerobico",      "valor": cfg["estequiometria_biomasa"]["Yx_s_aerobico"], "unidad": "g CDM/g glucosa", "fuente": "Roels 1980"})
    bio_rows.append({"parametro": "Umbral µ overflow",  "valor": cfg["metabolitos_secundarios"]["acetato"]["mu_umbral_h"], "unidad": "h⁻¹", "fuente": "Luli & Strohl 1990"})
    bio_rows.append({"parametro": "Osm umbral suave",   "valor": cfg["osmolaridad"]["umbral_suave_mOsm_kg"], "unidad": "mOsm/kg", "fuente": "Cayley et al 1991"})
    bio_rows.append({"parametro": "Osm umbral fuerte",  "valor": cfg["osmolaridad"]["umbral_fuerte_mOsm_kg"], "unidad": "mOsm/kg", "fuente": "Cayley et al 1991"})
    bio_rows.append({"parametro": "Osm cap letal",      "valor": cfg["osmolaridad"]["cap_letal_mOsm_kg"], "unidad": "mOsm/kg", "fuente": "Shabala et al 2009"})
    df_bio = pd.DataFrame(bio_rows)

    Path(output_path).parent.mkdir(exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_media.to_excel(  writer, sheet_name="suggested_media",   index=False)
        df_resumen.to_excel(writer, sheet_name="resumen",           index=False)
        df_adv.to_excel(    writer, sheet_name="advertencias_bio",  index=False)
        filter_df.to_excel( writer, sheet_name="filter_report",     index=False)
        df_bio.to_excel(    writer, sheet_name="bio_constraints",   index=False)
        _format_excel(writer)


def _format_excel(writer):
    COLORES = {
        "suggested_media":  "1F4E79",
        "resumen":          "375623",
        "advertencias_bio": "7B2D2D",
        "filter_report":    "412402",
        "bio_constraints":  "26215C",
    }
    for sheet_name, hex_color in COLORES.items():
        if sheet_name not in writer.sheets:
            continue
        ws   = writer.sheets[sheet_name]
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        for cell in ws[1]:
            cell.font      = Font(color="FFFFFF", bold=True, size=10)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

        if sheet_name == "suggested_media":
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                n_adv_cell = next(
                    (c for c in row if c.column == _col_index(ws, "n_advertencias")), None)
                if n_adv_cell and (n_adv_cell.value or 0) > 0:
                    for c in row:
                        c.fill = PatternFill(start_color="FAEEDA", end_color="FAEEDA",
                                              fill_type="solid")
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 35)

    tab_colors = {
        "suggested_media": "1F4E79", "resumen": "375623",
        "advertencias_bio": "7B2D2D", "bio_constraints": "26215C"
    }
    for name, color in tab_colors.items():
        if name in writer.sheets:
            writer.sheets[name].sheet_properties.tabColor = color


def _col_index(ws, col_name: str) -> Optional[int]:
    for cell in ws[1]:
        if cell.value == col_name:
            return cell.column
    return None


def run_predict(data_path: str, input_path: str, output_path: str):
    """Entrena si es necesario, luego lanza el predictor."""
    from predictor import run_predictor, print_results

    model_path = Path(MODEL_PATH)
    if not model_path.exists():
        logger.info("Modelo no encontrado. Entrenando primero...")
        train_model(data_path)
    resultados = run_predictor(
        model_path  = str(model_path),
        input_path  = input_path,
        output_path = output_path,
        group_override = TARGET_GROUP,
    )
    print_results(resultados)
    print(f"\n  Resultados exportados: {output_path}\n")


def parse_args():
    p = argparse.ArgumentParser(
        description="Alchemist v1.0 — Formulador y predictor de medios para E.coli",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  # Formular los 10 mejores medios
  python main.py --mode optimize --data data.xlsx

  # Formular con mas iteraciones
  python main.py --mode optimize --data data.xlsx --n_trials 400 --n_top 10

  # Predecir crecimiento en un medio propuesto
  python main.py --mode predict --data data.xlsx --input mi_medio.xlsx

  # Predecir con salida personalizada
  python main.py --mode predict --data data.xlsx --input mi_medio.xlsx --output resultado.xlsx
        """
    )
    p.add_argument("--mode",      required=True, choices=["optimize","predict"],
                   help="'optimize': formula los 10 mejores medios. 'predict': evalua un medio del usuario.")
    p.add_argument("--data",      required=True,
                   help="Excel con el dataset de entrenamiento (experiment_data, growth_curve, culture_conditions).")
    p.add_argument("--input",     default=None,
                   help="[Solo --mode predict] Excel con el medio a evaluar.")
    p.add_argument("--output",    default=None,
                   help="Archivo Excel de salida. Default: results/alchemist_results.xlsx (optimize) o predictor_results.xlsx (predict).")
    p.add_argument("--n_trials",  type=int, default=200,
                   help="Numero de iteraciones del optimizador (default: 200).")
    p.add_argument("--n_top",     type=int, default=10,
                   help="Numero de mejores medios a devolver (default: 10).")
    p.add_argument("--log_level", default="INFO",
                   choices=["DEBUG","INFO","WARNING","ERROR"])
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    _setup_logging(args.log_level)

    _banner("ALCHEMIST v1.0")
    print(f"  Modo:        {args.mode.upper()}")
    print(f"  Dataset:     {args.data}")
    if args.mode == "predict":
        print(f"  Medio:       {args.input}")
    print(f"  Restricciones biologicas: config/bio_config_ecoli.json")
    print()

    try:
        if args.mode == "optimize":
            output = args.output or RESULTS_EXCEL
            run_optimize(args.data, args.n_trials, args.n_top)

        elif args.mode == "predict":
            if not args.input:
                print("ERROR: --mode predict requiere --input <excel_del_usuario>")
                sys.exit(1)
            output = args.output or "predictor_results.xlsx"
            run_predict(args.data, args.input, output)

    except FileNotFoundError as e:
        print(f"\n  ERROR: {e}\n")
        sys.exit(1)
    except Exception as e:
        logger.exception(e)
        print(f"\n  ERROR inesperado: {e}\n")
        sys.exit(2)
