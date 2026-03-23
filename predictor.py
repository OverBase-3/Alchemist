import argparse
import logging
import sys
import math
from pathlib import Path

import numpy as np
import pandas as pd
import joblib

from bio_constraints import (
    get_config,
    enforce_env_bounds,
    stoichiometric_min_nitrogen,
    calc_available_nitrogen,
    calc_available_carbon,
    calculate_osmolarity,
    cdm_ceiling,
    check_missing_micronutrients,
    cross_penalty_mu_cdm,
    osmolarity_penalty,
    stoichiometric_penalty,
)
from bio_groups import normalize_cell_type

logger = logging.getLogger(__name__)

class ModelBundle:

    def __init__(self, model_path: str):
        self.path = Path(model_path)
        if not self.path.exists():
            raise FileNotFoundError(
                f"Modelo no encontrado: {self.path}\n"
                "Ejecuta primero: python main.py --mode optimize --data data.xlsx"
            )
        bundle = joblib.load(self.path)
        self.trainers            = bundle["trainers"]
        self.feature_engineers   = bundle["feature_engineers"]
        self.exp_data_by_group   = bundle["exp_data_by_group"]

    def available_groups(self):
        return list(self.trainers.keys())

    def predict(self, medium_df: pd.DataFrame,
                conditions_df: pd.DataFrame,
                group: str) -> pd.DataFrame:
        if group not in self.trainers:
            raise ValueError(
                f"Grupo '{group}' no entrenado. Disponibles: {self.available_groups()}"
            )
        fe      = self.feature_engineers[group]
        trainer = self.trainers[group]
        X       = fe.transform(medium_df, conditions_df)
        preds   = trainer.predict(X)
        if "max_cell_density" in preds.columns:
            zero_cdm = preds["max_cell_density"] < 0.01
            if zero_cdm.any() and "RandomForest" in trainer.results:
                rf = trainer.results["RandomForest"]["model"]
                rf_pred = rf.predict(X)
                targets = list(trainer.targets)
                if "max_cell_density" in targets:
                    idx = targets.index("max_cell_density")
                    rf_mcd = rf_pred[:, idx] if rf_pred.ndim > 1 else rf_pred
                    preds.loc[zero_cdm, "max_cell_density"] = np.clip(
                        rf_mcd[zero_cdm.values], 0, None
                    )
        return preds

def load_user_excel(filepath: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    xl     = pd.ExcelFile(filepath)
    sheets = xl.sheet_names

    if "experiment_data" not in sheets:
        raise ValueError(
            f"El archivo '{filepath}' no tiene hoja 'experiment_data'.\n"
            "Columnas requeridas: medium_id, cell_type, component, concentration, unit."
        )

    exp = xl.parse("experiment_data")
    exp["component"] = (exp["component"].astype(str).str.strip().str.lower()
                         .str.replace(r"[\s\-]+", "_", regex=True)
                         .str.replace(r"[^\w]", "", regex=True))
    exp["cell_type"] = exp["cell_type"].astype(str).str.strip()
    exp["medium_id"] = exp["medium_id"].astype(str).str.strip()

    for col in ["growth_rate", "max_cell_density", "doubling_time"]:
        if col not in exp.columns:
            exp[col] = 0.0
        else:
            exp[col] = pd.to_numeric(exp[col], errors="coerce").fillna(0.0)
    if "culture_conditions" in sheets:
        cc = xl.parse("culture_conditions")
        if "carriers" in cc.columns:
            cc["carriers"] = (cc["carriers"].astype(str).str.lower()
                               .map({"yes": 1.0, "no": 0.0}).fillna(0.0))
    else:
        media_ids = exp["medium_id"].unique()
        cc = pd.DataFrame([{
            "medium_id":      m,
            "temperature_C":  37.0,
            "pH":             7.0,
            "carriers":       0.0,
            "agitation_rpm":  200.0,
        } for m in media_ids])
        logger.warning("Hoja 'culture_conditions' no encontrada. Usando valores por defecto.")

    return exp, cc

def evaluate_user_medium(medium_id: str,
                          composition: dict,
                          conditions: dict,
                          predicted_mu: float,
                          predicted_cdm: float,
                          group: str) -> list:

    advertencias = []

    cfg_env = get_config(group)["condiciones_entorno"]
    for var in ["temperature_C", "pH", "agitation_rpm"]:
        if var not in conditions:
            continue
        val = conditions[var]
        lo  = cfg_env[var]["min"]
        hi  = cfg_env[var]["max"]
        if val < lo:
            advertencias.append({
                "capa": "Capa 1 — Condiciones de entorno",
                "variable": var,
                "valor": val,
                "rango_permitido": f"[{lo} – {hi}]",
                "mensaje": cfg_env[var].get("advertencia_min", f"{var} demasiado bajo."),
            })
        elif val > hi:
            advertencias.append({
                "capa": "Capa 1 — Condiciones de entorno",
                "variable": var,
                "valor": val,
                "rango_permitido": f"[{lo} – {hi}]",
                "mensaje": cfg_env[var].get("advertencia_max", f"{var} demasiado alto."),
            })

    if predicted_cdm > 0:
        N_avail = calc_available_nitrogen(composition, group)
        N_req   = stoichiometric_min_nitrogen(predicted_cdm, group)
        N_def   = max(0.0, (N_req - N_avail) / max(N_req, 1e-9))
        cfg_bio = get_config(group)["estequiometria_biomasa"]
        if N_def > cfg_bio["N_deficit_tolerancia"]:
            advertencias.append({
                "capa": "Capa 2 — Balance de nitrogeno (Pirt)",
                "N_disponible_g_L": round(N_avail, 4),
                "N_requerido_g_L":  round(N_req, 4),
                "deficit_pct":      round(N_def * 100, 1),
                "mensaje": (f"Deficit de N = {N_def*100:.1f}%. "
                             + cfg_bio["advertencia_deficit_N"]),
            })

    faltantes = check_missing_micronutrients(composition, group)
    advertencias.extend(faltantes)

    _, adv4 = cross_penalty_mu_cdm(predicted_mu, predicted_cdm, group)
    advertencias.extend(adv4)

    _, adv5 = osmolarity_penalty(composition, group)
    advertencias.extend(adv5)

    return advertencias


def run_predictor(model_path: str, input_path: str,
                  output_path: str = "predictor_results.xlsx",
                  group_override: str = None) -> list:

    bundle = ModelBundle(model_path)
    exp_df, cc_df = load_user_excel(input_path)

    resultados = []

    for medium_id in exp_df["medium_id"].unique():
        sub      = exp_df[exp_df["medium_id"] == medium_id].copy()
        cell_raw = sub["cell_type"].iloc[0]
        group    = group_override or normalize_cell_type(cell_raw)

        if group not in bundle.available_groups():
            resultados.append({
                "medium_id": medium_id,
                "cell_type": group,
                "error": (f"Grupo '{group}' no entrenado en el modelo. "
                          f"Grupos disponibles: {bundle.available_groups()}"),
                "mu_h":   None, "td_h": None, "CDM_g_L": None,
                "advertencias": [],
            })
            continue

        cc_sub = cc_df[cc_df["medium_id"] == medium_id]
        if cc_sub.empty:
            conds = {"temperature_C": 37.0, "pH": 7.0,
                     "carriers": 0.0, "agitation_rpm": 200.0}
        else:
            conds = cc_sub.iloc[0].to_dict()

        try:
            preds = bundle.predict(sub, cc_df, group)
            mu_pred  = float(np.clip(preds["growth_rate"].iloc[0], 0, None))
            cdm_pred = float(np.clip(preds["max_cell_density"].iloc[0], 0, None))
            td_pred  = float(math.log(2) / mu_pred) if mu_pred > 1e-6 else float("nan")
        except Exception as e:
            resultados.append({
                "medium_id": medium_id,
                "cell_type": group,
                "error": f"Error en prediccion ML: {e}",
                "mu_h": None, "td_h": None, "CDM_g_L": None,
                "advertencias": [],
            })
            continue

        composition = dict(zip(sub["component"], sub["concentration"].astype(float)))

        advertencias = evaluate_user_medium(
            medium_id, composition, conds, mu_pred, cdm_pred, group
        )

        resultados.append({
            "medium_id":   medium_id,
            "cell_type":   group,
            "mu_h":        round(mu_pred, 5),
            "td_h":        round(td_pred, 3) if not math.isnan(td_pred) else None,
            "CDM_g_L":     round(cdm_pred, 4),
            "n_advertencias": len(advertencias),
            "advertencias":   advertencias,
        })

    _export_predictor_excel(resultados, output_path)
    return resultados


def _export_predictor_excel(resultados: list, output_path: str):
    """Exporta los resultados del predictor a Excel."""
    from openpyxl.styles import Font, PatternFill, Alignment

    rows_pred = []
    rows_adv  = []

    for r in resultados:
        rows_pred.append({
            "medium_id":      r["medium_id"],
            "cell_type":      r["cell_type"],
            "µ_predicho_h":   r["mu_h"],
            "td_predicho_h":  r["td_h"],
            "CDM_predicha_g_L": r["CDM_g_L"],
            "n_advertencias": r.get("n_advertencias", 0),
            "error":          r.get("error", ""),
        })
        for adv in r.get("advertencias", []):
            rows_adv.append({
                "medium_id": r["medium_id"],
                "capa":      adv.get("capa", ""),
                "variable":  adv.get("variable", adv.get("componente", "")),
                "mensaje":   adv.get("mensaje", ""),
                "detalle":   str({k: v for k, v in adv.items()
                                  if k not in ("capa","mensaje","variable","componente")}),
            })

    df_pred = pd.DataFrame(rows_pred)
    df_adv  = pd.DataFrame(rows_adv) if rows_adv else pd.DataFrame(
        columns=["medium_id","capa","variable","mensaje","detalle"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_pred.to_excel(writer, sheet_name="predicciones",   index=False)
        df_adv.to_excel(writer,  sheet_name="advertencias",   index=False)

        ws = writer.sheets["predicciones"]
        hdr = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.fill = hdr
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            n_adv = row[5].value or 0
            color = "FAEEDA" if n_adv > 0 else "FFFFFF"
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color,
                                         fill_type="solid")
                cell.font = Font(size=9)
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 30)

        ws2 = writer.sheets["advertencias"]
        hdr2 = PatternFill(start_color="7B2D2D", end_color="7B2D2D", fill_type="solid")
        for cell in ws2[1]:
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.fill = hdr2
        for col in ws2.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws2.column_dimensions[col[0].column_letter].width = min(mx + 3, 60)

    logger.info(f"Resultados del predictor exportados: {output_path}")


def print_results(resultados: list):
    SEP = "═" * 62

    for r in resultados:
        print(f"\n{SEP}")
        print(f"  Medio: {r['medium_id']}  |  Tipo celular: {r['cell_type']}")
        print(SEP)

        if r.get("error"):
            print(f"  ✗ ERROR: {r['error']}")
            continue

        print(f"  µ (tasa de crecimiento):     {r['mu_h']:.5f} h⁻¹")
        td = r['td_h']
        print(f"  td (tiempo de duplicacion):  {td:.2f} h" if td else
              f"  td (tiempo de duplicacion):  N/D")
        print(f"  CDM (densidad celular max):  {r['CDM_g_L']:.4f} g/L")

        advs = r.get("advertencias", [])
        if advs:
            print(f"\n    {len(advs)} ADVERTENCIA(S) BIOLOGICA(S):")
            print("  " + "─" * 56)
            for i, adv in enumerate(advs, 1):
                print(f"\n  [{i}] {adv.get('capa','')}")
                if "variable" in adv:
                    print(f"      Variable:      {adv['variable']}")
                    print(f"      Valor actual:  {adv.get('valor','')}")
                    print(f"      Rango valido:  {adv.get('rango_permitido','')}")
                if "N_disponible_g_L" in adv:
                    print(f"      N disponible:  {adv['N_disponible_g_L']} g/L")
                    print(f"      N requerido:   {adv['N_requerido_g_L']} g/L")
                    print(f"      Deficit:       {adv['deficit_pct']}%")
                if "osm_total" in adv:
                    print(f"      Osmolaridad:   {adv['osm_total']:.0f} mOsm/kg")
                if "mu_predicho" in adv:
                    print(f"      µ predicho:    {adv['mu_predicho']} h⁻¹")
                    print(f"      CDM techo:     {adv.get('cdm_ceiling','?')} g/L")
                msg = adv.get("mensaje", "")
                if msg:
                    words = msg.split()
                    line = "      Detalle: "
                    for w in words:
                        if len(line) + len(w) > 78:
                            print(line)
                            line = "               " + w + " "
                        else:
                            line += w + " "
                    print(line)
        else:
            print(f"\n  ✓ Sin advertencias biologicas.")

    print(f"\n{SEP}")


def parse_args():
    p = argparse.ArgumentParser(
        description="Alchemist v1.0 — Predictor de crecimiento",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python predictor.py --model results/alchemist_model.pkl --input mi_medio.xlsx
  python predictor.py --model results/alchemist_model.pkl --input mi_medio.xlsx --output resultados.xlsx
  python predictor.py --model results/alchemist_model.pkl --input mi_medio.xlsx --cell_type E.coli
        """
    )
    p.add_argument("--model",     required=True, help="Ruta al modelo .pkl entrenado")
    p.add_argument("--input",     required=True, help="Excel con el medio a evaluar")
    p.add_argument("--output",    default="predictor_results.xlsx",
                   help="Archivo de salida Excel (default: predictor_results.xlsx)")
    p.add_argument("--cell_type", default=None,
                   help="Forzar tipo celular (e.g. E.coli). Si no se indica, se lee del Excel.")
    p.add_argument("--log_level", default="WARNING",
                   choices=["DEBUG", "INFO", "WARNING"])
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    print("\n" + "═" * 62)
    print("  ALCHEMIST v1.0 — Predictor de crecimiento")
    print("═" * 62)

    try:
        resultados = run_predictor(
            model_path    = args.model,
            input_path    = args.input,
            output_path   = args.output,
            group_override= args.cell_type,
        )
        print_results(resultados)
        print(f"\n  Resultados exportados: {args.output}\n")
        sys.exit(0)
    except FileNotFoundError as e:
        print(f"\n  ERROR: {e}\n")
        sys.exit(1)
    except Exception as e:
        logger.exception(e)
        print(f"\n  ERROR inesperado: {e}\n")
        sys.exit(2)
